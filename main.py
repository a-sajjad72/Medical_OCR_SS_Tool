import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import os
import logging
import threading
import tempfile
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import subprocess
from utils import resource_path, get_tessbin_path, get_tessdata_path
import traceback
import sys
import site
import time
import pyautogui
from screenshot import select_region

if site.USER_SITE is None:
    # Set a fallback value.
    site.USER_SITE = resource_path(".")


os.environ["PADDLE_OCR_BASE_DIR"] = resource_path("./models/paddleocr")
os.environ["EASYOCR_MODULE_PATH"] = resource_path("./models/easyocr")
os.environ["TESSDATA_PREFIX"] = get_tessdata_path()

# Set up logging
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] [%(levelname)8s] %(message)s')
logger = logging.getLogger(__name__)

class OCRApp:
    def __init__(self, root):
        self.root = root
        self.root.title("OCR to Excel Converter")
        self.root.geometry("1280x720")
        # Set application icon to 'icon.png'
        self.icon_image = tk.PhotoImage(file=resource_path('icons/icon.png'))
        self.root.iconphoto(False, self.icon_image)
        self.style = ttk.Style('cosmo')  # Use a modern theme
        self.ocr_engine = tk.StringVar()
        self.ocr_engine.set("PaddleOCR")

        # Confidence thresholds
        self.green_threshold = tk.IntVar(value=97)
        self.yellow_threshold = tk.IntVar(value=92)

        # Output directory
        self.output_directory = None

        self.setup_ui()

        self.ocr_models = {
            "PaddleOCR": None,
            "Tesseract": None,
            "EasyOCR": None
        }
        self.loading_status = tk.StringVar()
        self.setup_loading_screen()
        self.root.after(100, self.preload_engines)  # Non-blocking preload

    def setup_loading_screen(self):
        self.loading_frame = ttk.Frame(self.root)
        self.loading_frame.pack(fill=tk.BOTH, expand=True)
        
        self.loading_label = ttk.Label(
            self.loading_frame, 
            textvariable=self.loading_status,
            font=('Helvetica', 14)
        )
        self.loading_label.pack(pady=20)
        
        self.loading_progress = ttk.Progressbar(
            self.loading_frame, 
            mode='indeterminate'
        )
        self.loading_progress.pack(pady=10)
        
        self.loading_progress.start()

    def preload_engines(self):
        """Non-blocking background preloading"""
        def _preload():
            engines = list(self.ocr_models.keys())
            for engine in engines:
                self.loading_status.set(f"Preloading {engine}...")
                try:
                    if engine == "PaddleOCR":
                        from OCR_Modules.paddleOCR import initialize_ocr_SLANet_LCNetV2
                        self.ocr_models[engine] = initialize_ocr_SLANet_LCNetV2()
                    elif engine == "Tesseract":
                        from OCR_Modules.tesseractOCR import initialize_tesseract
                        self.ocr_models[engine] = initialize_tesseract(get_tessbin_path())
                    elif engine == "EasyOCR":
                        from OCR_Modules.easyOCR import initialize_easyocr
                        self.ocr_models[engine] = initialize_easyocr()
                except Exception as e:
                    logger.error(f"Error preloading {engine}: {str(e)}")
                
            self.loading_frame.pack_forget()
        
        threading.Thread(target=_preload).start()

    def setup_ui(self):
        # Main frame
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Center frame for initial view
        self.center_frame = ttk.Frame(self.main_frame)
        self.center_frame.pack(expand=True)

        # Logo
        logo_img = Image.open(resource_path("icons/icon.png"))
        logo_img = logo_img.resize((100, 100), Image.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(logo_img)
        self.logo_label = ttk.Label(self.center_frame, image=self.logo_photo)
        self.logo_label.pack(pady=(20, 10))

        # Output Directory Selection
        folder_icon = Image.open(resource_path("icons/folder.png"))
        folder_icon = folder_icon.resize((20, 20), Image.LANCZOS)
        self.folder_icon_photo = ImageTk.PhotoImage(folder_icon)
        output_dir_button = ttk.Button(self.center_frame, text="Select Output Directory", 
                                       image=self.folder_icon_photo, compound=tk.LEFT,
                                       command=self.select_output_directory)
        output_dir_button.pack(pady=(10, 5))

        # OCR Engine Selection
        ocr_label = ttk.Label(self.center_frame, text="Select OCR Engine:")
        ocr_label.pack(pady=(10, 5))
        ocr_dropdown = ttk.Combobox(self.center_frame, textvariable=self.ocr_engine, state="readonly", width=30)
        ocr_dropdown['values'] = ('PaddleOCR', 'Tesseract', 'EasyOCR')
        ocr_dropdown.pack(pady=(0, 20))

        # Confidence Thresholds
        thresholds_frame = ttk.Frame(self.center_frame)
        thresholds_frame.pack(pady=(10, 20))

        options = [str(i) for i in range(80, 101)]  # Values from 80 to 100

        green_label = ttk.Label(thresholds_frame, text="High Confidence Threshold (Green) (> %):")
        green_label.grid(row=0, column=0, padx=5, pady=5, sticky='e')
        green_dropdown = ttk.Combobox(thresholds_frame, textvariable=self.green_threshold, values=options, state='readonly', width=5)
        green_dropdown.grid(row=0, column=1, padx=5, pady=5)
        green_dropdown.bind('<<ComboboxSelected>>', self.update_thresholds)

        yellow_label = ttk.Label(thresholds_frame, text="Medium Confidence Threshold (Yellow) (> %):")
        yellow_label.grid(row=1, column=0, padx=5, pady=5, sticky='e')
        yellow_dropdown = ttk.Combobox(thresholds_frame, textvariable=self.yellow_threshold, values=options, state='readonly', width=5)
        yellow_dropdown.grid(row=1, column=1, padx=5, pady=5)
        yellow_dropdown.bind('<<ComboboxSelected>>', self.update_thresholds)

        # Upload Button
        upload_icon = Image.open(resource_path("icons/upload.png"))
        upload_icon = upload_icon.resize((20, 20), Image.LANCZOS)
        self.upload_icon_photo = ImageTk.PhotoImage(upload_icon)
        self.upload_button = ttk.Button(self.center_frame, text="Upload Image", 
                                        image=self.upload_icon_photo, compound=tk.LEFT,
                                        command=self.select_image, width=20)
        self.upload_button.pack(pady=(0, 10))

        # Screenshot Button
        screenshot_icon = Image.open(resource_path("icons/screenshot.png"))
        screenshot_icon = screenshot_icon.resize((20, 20), Image.LANCZOS)
        self.screenshot_icon_photo = ImageTk.PhotoImage(screenshot_icon)
        self.screenshot_button = ttk.Button(self.center_frame, text="Screenshot", 
                                            image=self.screenshot_icon_photo, compound=tk.LEFT,
                                            command=self.take_screenshot, width=20)
        self.screenshot_button.pack(pady=(0, 20))

        # Status Label
        self.status_label = ttk.Label(self.center_frame, text="")
        self.status_label.pack(pady=(0, 10))

        # Prepare frames for results
        self.top_frame = ttk.Frame(self.main_frame)
        self.left_frame = ttk.Frame(self.main_frame)
        self.middle_frame = ttk.Frame(self.main_frame)
        self.right_frame = ttk.Frame(self.main_frame)  # For the sidebar

        # Modify progress bar initialization
        self.progress_bar = ttk.Progressbar(
            self.center_frame, 
            mode='indeterminate',
            length=300  # Explicit length for better macOS visibility
        )

    def select_output_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_directory = directory
            logger.info(f"Output directory set to: {self.output_directory}")
            self.status_label.config(text=f"Output directory set to: {self.output_directory}")

    def update_thresholds(self, event=None):
        # Ensure that green_threshold is always greater than yellow_threshold
        if self.green_threshold.get() < self.yellow_threshold.get():
            self.yellow_threshold.set(self.green_threshold.get())

    def reset_ui(self):
        if sys.platform == "darwin":
            self.root.update_idletasks()
            self.root.update()

        # Hide the top, left, middle, and right frames
        self.top_frame.pack_forget()
        self.left_frame.pack_forget()
        self.middle_frame.pack_forget()
        self.right_frame.pack_forget()

        # Clear images from frames
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.middle_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Show the center frame
        self.center_frame.pack(expand=True)

        # Reset status label
        self.status_label.config(text="")

        # Remove progress bar
        self.progress_bar.pack_forget()

    def select_image(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Image files", ("*.png", "*.jpg", "*.jpeg", "*.bmp", "*.tiff"))]
        )
        if file_path:
            logger.info(f"Selected image: {file_path}")
            self.is_screenshot = False  # Indicate that this is not a screenshot
            self.reset_ui()  # Reset the UI before processing a new image
            self.process_image(file_path)

    def take_screenshot(self):
        # Minimize the root window
        root.withdraw()
        root.update()  # Force Tkinter to process the hide request
        time.sleep(0.5)  # Give the OS time to hide the window (adjust as needed)

        # Now take the (initial) screenshot 
        screenshot = pyautogui.screenshot().convert("RGB")
        region = select_region(root, screenshot)

        # Restore the root window
        self.root.deiconify()

        if region:
            x1, y1, x2, y2 = region
            snip_img = screenshot.crop((x1, y1, x2, y2))

        # Determine the output directory
        if self.output_directory:
            output_dir = self.output_directory
        else:
            # For screenshots, default to Desktop
            output_dir = os.path.join(os.path.expanduser("~"), "Desktop")

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        # Save the screenshot image
        base_filename = "screenshot"
        screenshot_path = os.path.join(output_dir, base_filename + ".png")
        snip_img.save(screenshot_path)

        # Reset the UI before processing
        self.reset_ui()

        # Process the image
        self.is_screenshot = True  # Indicate that this is a screenshot
        self.process_image(screenshot_path)

    def process_image(self, file_path):
        self.show_progress("Initializing OCR Engine...")
        # Use main thread for macOS UI operations
        if sys.platform == "darwin":
            self.root.after(0, self._process_image_thread, file_path)
        else:
            processing_thread = threading.Thread(
                target=self._process_image_thread, args=(file_path,)
            )
            processing_thread.start()

    def show_progress(self, message):
        self.status_label.config(text=message)
        if sys.platform == "darwin":
            self.root.update_idletasks()
        self.progress_bar.pack(pady=(0, 10))
        self.progress_bar.start()
        print(f"called show_progress with {message}")

    def hide_progress(self):
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.status_label.config(text="")
        if sys.platform == "darwin":
            self.root.update_idletasks()
        print("called hide_progress")

    def _process_image_thread(self, file_path):
        try:
            ocr_engine = self.ocr_engine.get()

            # Check if engine needs initialization
            if not self.ocr_models.get(ocr_engine):
                self.show_progress(f"Loading {ocr_engine}...")
                # Lazy load the engine if not preloaded
                if ocr_engine == "PaddleOCR":
                    from OCR_Modules.paddleOCR import initialize_ocr_SLANet_LCNetV2
                    self.ocr_models[ocr_engine] = initialize_ocr_SLANet_LCNetV2()
                elif ocr_engine == "Tesseract":
                    from OCR_Modules.tesseractOCR import initialize_tesseract
                    self.ocr_models[ocr_engine] = initialize_tesseract(get_tessbin_path())
                elif ocr_engine == "EasyOCR":
                    from OCR_Modules.easyOCR import initialize_easyocr
                    self.ocr_models[ocr_engine] = initialize_easyocr()
                self.hide_progress()

            self.show_progress("Processing image...")    

            # Show progress bar in the center frame
            self.progress_bar.pack(pady=(0, 10))
            self.progress_bar.start()

            if ocr_engine == "PaddleOCR":
                self.process_with_paddleocr(file_path)
            elif ocr_engine == "Tesseract":
                self.process_with_tesseract(file_path)
            elif ocr_engine == "EasyOCR":
                self.process_with_easyocr(file_path)
            else:
                raise ValueError("Please select an OCR engine.")
        except Exception as e:
            logger.error(f"Error processing image: {str(e)}")
            self.status_label.config(text=f"Error: {str(e)}\nPlease try a different image or OCR engine.")
        finally:
            self.hide_progress()

    def process_with_paddleocr(self, file_path):
        from OCR_Modules.paddleOCR import (
            process_image as paddle_process_image,
            group_into_rows as paddle_group_into_rows,
            save_as_xlsx as paddle_save_as_xlsx,
            draw_bounding_boxes as paddle_draw_bounding_boxes,
        )

        data = paddle_process_image(file_path, self.ocr_models["PaddleOCR"])
        rows = paddle_group_into_rows(data)

        if not rows:
            raise ValueError("No data extracted from image.")

        # Determine the output directory
        if self.output_directory:
            output_dir = self.output_directory
        else:
            if self.is_screenshot:
                # For screenshots, default to Desktop
                output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            else:
                # For uploaded images, use the same directory as the image
                output_dir = os.path.dirname(file_path)
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        # Create the output filenames
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
        output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

        # Get thresholds from dropdowns
        green_thresh = self.green_threshold.get() / 100.0
        yellow_thresh = self.yellow_threshold.get() / 100.0

        paddle_save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)

        paddle_draw_bounding_boxes(file_path, data, output_image_path)

        self.status_label.config(text=f"Excel file saved: {output_xlsx}")
        self.display_results(output_image_path, output_xlsx)

    def process_with_tesseract(self, file_path):
        try:
            from OCR_Modules.tesseractOCR import (
                process_image as tesseract_process_image,
                group_into_rows as tesseract_group_into_rows,
                save_as_xlsx as tesseract_save_as_xlsx,
                draw_bounding_boxes as tesseract_draw_bounding_boxes,
            )

            data = tesseract_process_image(file_path, self.ocr_models["Tesseract"])

            if not data:
                raise ValueError("No data extracted from image.")

            rows = tesseract_group_into_rows(data)

            if not rows:
                raise ValueError("No rows extracted from data.")

            # Determine the output directory
            if self.output_directory:
                output_dir = self.output_directory
            else:
                if self.is_screenshot:
                    # For screenshots, default to Desktop
                    output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                else:
                    # For uploaded images, use the same directory as the image
                    output_dir = os.path.dirname(file_path)
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            # Create the output filenames
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
            output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

            # Get thresholds from dropdowns
            green_thresh = self.green_threshold.get() / 100.0
            yellow_thresh = self.yellow_threshold.get() / 100.0

            tesseract_save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)

            tesseract_draw_bounding_boxes(file_path, data, output_image_path)

            self.status_label.config(text=f"Excel file saved: {output_xlsx}")
            self.display_results(output_image_path, output_xlsx)
        except ValueError as ve:
            logger.error(f"Tesseract processing error: {str(ve)}")
            self.status_label.config(text=f"Error: {str(ve)}\nPlease try a different image or OCR engine.")
        except Exception as e:
            logger.error(f"Unexpected error in Tesseract processing: {str(e)}")
            self.status_label.config(text=f"Unexpected error: {str(e)}\nPlease try a different image or OCR engine.")

    def process_with_easyocr(self, file_path):
        try:
            from OCR_Modules.easyOCR import (
                process_image as easyocr_process_image,
                group_into_rows as easyocr_group_into_rows,
                save_as_xlsx as easyocr_save_as_xlsx,
                draw_bounding_boxes as easyocr_draw_bounding_boxes,
            )

            data = easyocr_process_image(file_path, self.ocr_models["EasyOCR"])

            if not data:
                raise ValueError("No data extracted from image.")

            rows = easyocr_group_into_rows(data)

            if not rows:
                raise ValueError("No rows extracted from data.")

            # Determine the output directory
            if self.output_directory:
                output_dir = self.output_directory
            else:
                if self.is_screenshot:
                    # For screenshots, default to Desktop
                    output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                else:
                    # For uploaded images, use the same directory as the image
                    output_dir = os.path.dirname(file_path)
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            # Create the output filenames
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
            output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

            # Get thresholds from dropdowns
            green_thresh = self.green_threshold.get() / 100.0
            yellow_thresh = self.yellow_threshold.get() / 100.0

            easyocr_save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)

            easyocr_draw_bounding_boxes(file_path, data, output_image_path)

            self.status_label.config(text=f"Excel file saved: {output_xlsx}")
            self.display_results(output_image_path, output_xlsx)
        except ValueError as ve:
            logger.error(f"EasyOCR processing error: {str(ve)}")
            self.status_label.config(text=f"Error: {str(ve)}\nPlease try a different image or OCR engine.")
        except Exception as e:
            logger.error(f"Unexpected error in EasyOCR processing: {str(e)}")
            self.status_label.config(text=f"Unexpected error: {str(e)}\nPlease try a different image or OCR engine.")

    def display_results(self, image_path, excel_path):
        if sys.platform == "darwin":
            self.root.after(0, self._safe_display_results, image_path, excel_path)
        else:
            self._safe_display_results(image_path, excel_path)

    def _safe_display_results(self, image_path, excel_path):
        self.reorganize_layout()

        # Clear previous images
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.middle_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Display image with bounding boxes
        self.display_image(image_path, self.left_frame)

        # Display Excel image
        excel_image_path = os.path.splitext(excel_path)[0] + "_excel_image.png"
        self.generate_excel_image(excel_path, excel_image_path)

        # Add padding to the middle frame
        padding_frame = ttk.Frame(self.middle_frame, padding=20)
        padding_frame.pack(fill=tk.BOTH, expand=True)
        self.display_image(excel_image_path, padding_frame)

        # Setup the sidebar
        self.setup_sidebar()

    def reorganize_layout(self):
        if sys.platform == "darwin" and not self.root.winfo_exists():
            return

        # Main thread check for macOS
        if (
            sys.platform == "darwin"
            and threading.current_thread() != threading.main_thread()
        ):
            self.root.after(0, self.reorganize_layout)
            return

        # Hide the center frame
        self.center_frame.pack_forget()

        # Clear the top_frame
        for widget in self.top_frame.winfo_children():
            widget.destroy()

        # Set up top frame
        self.top_frame.pack(side=tk.TOP, fill=tk.X, pady=(10, 0))

        # Create a frame inside top_frame to center widgets
        self.root.update_idletasks()  # Refresh window state
        top_inner_frame = ttk.Frame(self.top_frame)
        top_inner_frame.pack(anchor='center')

        # Add small logo
        small_logo_img = Image.open(resource_path("icons/icon.png"))
        small_logo_img = small_logo_img.resize((50, 50), Image.LANCZOS)
        self.small_logo_photo = ImageTk.PhotoImage(small_logo_img)
        logo_label = ttk.Label(top_inner_frame, image=self.small_logo_photo)
        logo_label.pack(side=tk.LEFT, padx=(10, 5))

        ocr_label = ttk.Label(top_inner_frame, text="Select OCR Engine:")
        ocr_label.pack(side=tk.LEFT, padx=(10, 5))
        ocr_dropdown = ttk.Combobox(top_inner_frame, textvariable=self.ocr_engine, state="readonly", width=20)
        ocr_dropdown['values'] = ('PaddleOCR', 'Tesseract', 'EasyOCR')
        ocr_dropdown.pack(side=tk.LEFT, padx=(0, 10))
        upload_button = ttk.Button(top_inner_frame, text="Upload Image", command=self.select_image)
        upload_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Screenshot Button with Icon
        screenshot_icon = Image.open(resource_path(os.path.join('icons', 'screenshoticon.png')))
        screenshot_icon = screenshot_icon.resize((30, 30), Image.LANCZOS)
        self.screenshot_icon_photo = ImageTk.PhotoImage(screenshot_icon)
        screenshot_button = ttk.Button(top_inner_frame, image=self.screenshot_icon_photo, command=self.take_screenshot)
        screenshot_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Home Button with Icon
        home_icon = Image.open(resource_path(os.path.join('icons', 'home.png')))
        home_icon = home_icon.resize((30, 30), Image.LANCZOS)
        self.home_icon_photo = ImageTk.PhotoImage(home_icon)
        home_button = ttk.Button(top_inner_frame, image=self.home_icon_photo, command=self.reset_ui)
        home_button.pack(side=tk.LEFT, padx=(0, 10))

        # Set up frames
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.middle_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y)

    def display_image(self, image_path, panel):
        image = Image.open(image_path)

        # Create a canvas to display the image
        canvas = tk.Canvas(panel, bg='white')
        canvas.pack(fill=tk.BOTH, expand=True)

        # Function to resize the image when the panel size changes
        def resize_image(event):
            # Get the size of the canvas
            canvas_width = event.width
            canvas_height = event.height

            # Calculate the scaling factor
            width_ratio = canvas_width / image.width
            height_ratio = canvas_height / image.height
            scale_factor = min(width_ratio, height_ratio, 1.0)  # Do not upscale images

            new_width = int(image.width * scale_factor)
            new_height = int(image.height * scale_factor)

            resized_image = image.resize((new_width, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(resized_image)

            canvas.delete("all")
            canvas.create_image(canvas_width/2, canvas_height/2, image=photo, anchor='center')
            canvas.image = photo  # Keep a reference

        # Bind the resize event to the function
        canvas.bind("<Configure>", resize_image)

    def generate_excel_image(self, excel_path, output_image_path):
        from openpyxl import load_workbook
        from PIL import Image, ImageDraw, ImageFont

        wb = load_workbook(excel_path)
        ws = wb.active

        # Increase cell size and font size
        cell_width = 90
        cell_height = 30
        font_size = 80
        border_size = 40  # Increased border size

        max_col = ws.max_column
        max_row = ws.max_row
        image_width = cell_width * max_col + 2 * border_size
        image_height = cell_height * max_row + 2 * border_size

        image = Image.new('RGB', (image_width, image_height), 'white')
        draw = ImageDraw.Draw(image)
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except:
            font = ImageFont.load_default()

        for row in ws.iter_rows():
            for cell in row:
                col_idx = cell.column - 1
                row_idx = cell.row - 1
                x1 = col_idx * cell_width + border_size
                y1 = row_idx * cell_height + border_size
                x2 = x1 + cell_width
                y2 = y1 + cell_height

                # Draw cell background
                fill_color = 'FFFFFF'  # Default fill
                if cell.fill and cell.fill.fgColor:
                    if cell.fill.fgColor.type == 'rgb':
                        fill_color = cell.fill.fgColor.rgb[-6:]
                    elif cell.fill.fgColor.type == 'indexed':
                        fill_color = 'FFFFFF'  # Handle indexed colors as white

                draw.rectangle([x1, y1, x2, y2], fill=f'#{fill_color}', outline='black')

                # Draw cell text
                text = str(cell.value) if cell.value is not None else ''
                bbox = font.getbbox(text)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
                text_x = x1 + (cell_width - text_width) / 2
                text_y = y1 + (cell_height - text_height) / 2
                draw.text((text_x, text_y), text, fill='black', font=font)

        image.save(output_image_path)

    def setup_sidebar(self):
        # Sidebar content
        sidebar_frame = ttk.Frame(self.right_frame, padding=10)
        sidebar_frame.pack(fill=tk.Y, expand=False)

        # Accuracy Box
        accuracy_label = ttk.Label(sidebar_frame, text="Accuracy Thresholds", font=('Helvetica', 14, 'bold'))
        accuracy_label.pack(pady=(0, 10))

        # Icons for accuracy levels
        green_icon = Image.open(resource_path(os.path.join('icons', 'green_circle.png')))
        green_icon = green_icon.resize((20, 20), Image.LANCZOS)
        self.green_icon_photo = ImageTk.PhotoImage(green_icon)

        yellow_icon = Image.open(resource_path(os.path.join('icons', 'yellow_circle.png')))
        yellow_icon = yellow_icon.resize((20, 20), Image.LANCZOS)
        self.yellow_icon_photo = ImageTk.PhotoImage(yellow_icon)

        red_icon = Image.open(resource_path(os.path.join('icons', 'red_circle.png')))
        red_icon = red_icon.resize((20, 20), Image.LANCZOS)
        self.red_icon_photo = ImageTk.PhotoImage(red_icon)

        # Green Threshold Label
        green_frame = ttk.Frame(sidebar_frame)
        green_frame.pack(pady=5, anchor='w')
        green_label_icon = ttk.Label(green_frame, image=self.green_icon_photo)
        green_label_icon.pack(side=tk.LEFT)
        green_label_text = ttk.Label(green_frame, text=f"High Confidence (> {self.green_threshold.get()}%)")
        green_label_text.pack(side=tk.LEFT, padx=5)

        # Yellow Threshold Label
        yellow_frame = ttk.Frame(sidebar_frame)
        yellow_frame.pack(pady=5, anchor='w')
        yellow_label_icon = ttk.Label(yellow_frame, image=self.yellow_icon_photo)
        yellow_label_icon.pack(side=tk.LEFT)
        yellow_label_text = ttk.Label(yellow_frame, text=f"Medium Confidence (> {self.yellow_threshold.get()}%)")
        yellow_label_text.pack(side=tk.LEFT, padx=5)

        # Red Threshold Label
        red_frame = ttk.Frame(sidebar_frame)
        red_frame.pack(pady=5, anchor='w')
        red_label_icon = ttk.Label(red_frame, image=self.red_icon_photo)
        red_label_icon.pack(side=tk.LEFT)
        red_label_text = ttk.Label(red_frame, text=f"Low Confidence (≤ {self.yellow_threshold.get()}%)")
        red_label_text.pack(side=tk.LEFT, padx=5)

        # Divider
        separator = ttk.Separator(sidebar_frame, orient='horizontal')
        separator.pack(fill='x', pady=10)

        # Disclaimer Box
        disclaimer_frame = ttk.Frame(sidebar_frame)
        disclaimer_frame.pack(pady=(10, 10))

        # Info Icon
        info_icon = Image.open(resource_path(os.path.join('icons', 'info.png')))
        info_icon = info_icon.resize((20, 20), Image.LANCZOS)
        self.info_icon_photo = ImageTk.PhotoImage(info_icon)

        disclaimer_title_frame = ttk.Frame(disclaimer_frame)
        disclaimer_title_frame.pack(anchor='w')

        disclaimer_icon_label = ttk.Label(disclaimer_title_frame, image=self.info_icon_photo)
        disclaimer_icon_label.pack(side=tk.LEFT)

        disclaimer_label = ttk.Label(disclaimer_title_frame, text="Disclaimer", font=('Helvetica', 14, 'bold'))
        disclaimer_label.pack(side=tk.LEFT, padx=5)

        # Enhanced Disclaimer Text
        disclaimer_text = (
            "Note: When the input image contains a table with missing data in some rows or columns, "
            "the extracted data may not align perfectly in the output Excel file. "
            "Empty cells might cause subsequent data to shift positions, resulting in misaligned columns. "
            "Please review the Excel output carefully and adjust as needed."
        )
        disclaimer_message = ttk.Label(disclaimer_frame, text=disclaimer_text, wraplength=250, justify='left')
        disclaimer_message.pack(pady=5)

if __name__ == "__main__":
    root = ttk.Window(themename="cosmo")
    app = OCRApp(root)
    root.mainloop()